import pandas as pd
import os
import openpyxl
from openpyxl.styles import PatternFill

def process_excel_files(directory):
    if not os.path.exists(directory):
        print(f"Directory {directory} does not exist.")
        return

    for filename in os.listdir(directory):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(directory, filename)
            try:
                print(f"Processing {filename}")

                xls = pd.ExcelFile(file_path)
                sheets_to_keep = [sheet for sheet in xls.sheet_names if '_kab' in sheet or '_kec' in sheet]

                df_dict = {}

                for sheet in sheets_to_keep:
                    df = pd.read_excel(xls, sheet_name=sheet).fillna(0)
                    if 'prov' in df.columns:
                      df_filtered = df[df['prov'].isin([18])]
                      if not df_filtered.empty:
                          df_dict[sheet] = df_filtered
                    else:
                        df_dict[sheet] = df

                with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                    for sheet_name, df in df_dict.items():
                        df.to_excel(writer, sheet_name=sheet_name, index=False)

                wb = openpyxl.load_workbook(file_path)
                for sheet_name in df_dict:
                    if '_kec' in sheet_name:
                        ws = wb[sheet_name]
                        last_row = ws.max_row
                wb.save(file_path)
                print(f"Processed {filename}")
            except Exception as e:
                print(f"Failed to process {filename}: {e}")

directory = r'D:\Semester 7\Magang\Publikasi ST\Tanaman Perkebunan\backup 210824'
process_excel_files(directory)
